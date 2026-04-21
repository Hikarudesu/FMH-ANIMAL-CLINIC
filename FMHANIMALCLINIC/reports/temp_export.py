def export_analytics_excel(request):
    """Export analytics dashboard data to Excel (.xlsx) format divided by branches."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    user_branch = request.user.branch
    today = timezone.now().date()

    # ── Filters ─────────────────────────
    period = request.GET.get('period', 'monthly')

    if period == 'daily':
        date_from = today
        date_to = today
        period_label = today.strftime('%B %d, %Y')
    elif period == 'weekly':
        date_from = today - timedelta(days=today.weekday())
        date_to = date_from + timedelta(days=6)
        period_label = f"{date_from.strftime('%b %d')} – {date_to.strftime('%b %d, %Y')}"
    else:
        date_from = today.replace(day=1)
        _, last_day = monthrange(today.year, today.month)
        date_to = today.replace(day=last_day)
        period_label = today.strftime('%B %Y')

    branches = Branch.objects.filter(is_active=True).order_by('name')
    if request.user.is_module_branch_restricted('reports') and user_branch:
        branches = branches.filter(id=user_branch.id)

    wb = Workbook()

    # Styles
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='009688',
                              end_color='009688', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14)
    label_font = Font(name='Calibri', bold=True, size=11)
    value_font = Font(name='Calibri', size=11)
    money_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    is_first_sheet = True

    for current_branch in branches:
        # ── Gather data ───────────────────────────────────────────────────
        pets_qs = Pet.objects.filter(is_active=True)
        total_patients = pets_qs.count()
        new_patients_period = pets_qs.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).count()

        sales_qs = Sale.objects.filter(
            status=Sale.Status.COMPLETED,
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
            branch=current_branch
        )

        sales_agg = sales_qs.aggregate(
            gross_sales=Sum('subtotal'),
            net_sales=Sum('total'),
            transaction_count=Count('id'),
        )
        gross_sales = sales_agg['gross_sales'] or Decimal('0')
        net_sales = sales_agg['net_sales'] or Decimal('0')
        total_discount = Decimal('0')
        for sale in sales_qs:
            if sale.discount_percent > 0:
                total_discount += (sale.subtotal * sale.discount_percent /
                                   Decimal('100')).quantize(Decimal('0.01'))
        transaction_count = sales_agg['transaction_count'] or 0

        refund_qs = Refund.objects.filter(
            status=Refund.Status.COMPLETED,
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
            sale__branch=current_branch
        )
        total_refunds = refund_qs.aggregate(total=Sum('amount'))[
            'total'] or Decimal('0')
        net_sales -= total_refunds

        period_appts = Appointment.objects.filter(
            appointment_date__gte=date_from,
            appointment_date__lte=date_to,
            branch=current_branch
        ).exclude(status='CANCELLED')

        new_clients = period_appts.filter(is_returning_customer=False).count()
        returning_clients = period_appts.filter(is_returning_customer=True).count()

        # 12-month data
        months_data = []
        for i in range(11, -1, -1):
            m_date = today.replace(day=1)
            for _ in range(i):
                m_date = (m_date - timedelta(days=1)).replace(day=1)

            m_start = m_date
            _, m_last = monthrange(m_date.year, m_date.month)
            m_end = m_date.replace(day=m_last)

            appts_qs = Appointment.objects.filter(
                appointment_date__gte=m_start,
                appointment_date__lte=m_end,
                branch=current_branch
            ).exclude(status='CANCELLED')

            new_count = appts_qs.filter(is_returning_customer=False).count()
            returning_count = appts_qs.filter(is_returning_customer=True).count()

            m_sales_qs = Sale.objects.filter(
                status=Sale.Status.COMPLETED,
                created_at__date__gte=m_start,
                created_at__date__lte=m_end,
                branch=current_branch
            )

            m_sales_agg = m_sales_qs.aggregate(
                gross_sales=Sum('subtotal'),
                net_sales=Sum('total')
            )

            m_refund_qs = Refund.objects.filter(
                status=Refund.Status.COMPLETED,
                created_at__date__gte=m_start,
                created_at__date__lte=m_end,
                sale__branch=current_branch
            )
            m_refund_total = m_refund_qs.aggregate(total=Sum('amount'))[
                'total'] or Decimal('0')

            months_data.append({
                'month': m_date.strftime('%B %Y'),
                'new': new_count,
                'returning': returning_count,
                'gross': float(m_sales_agg['gross_sales'] or 0),
                'net': float((m_sales_agg['net_sales'] or Decimal('0')) - m_refund_total),
            })

        # ── Sheets for this branch ──────────────────────────────────────────────
        branch_safe = str(current_branch.name)[:15]

        if is_first_sheet:
            ws = wb.active
            is_first_sheet = False
        else:
            ws = wb.create_sheet()
            
        ws.title = f"Sum - {branch_safe}"
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

        ws.append(['FMH Animal Clinic — Analytics Report'])
        ws.merge_cells('A1:B1')
        ws['A1'].font = Font(name='Calibri', bold=True, size=16)

        ws.append([f'Period: {period_label}'])
        ws['A2'].font = Font(name='Calibri', italic=True, size=11)
        ws.append([f'Branch: {current_branch.name}'])
        ws.append([f'Generated: {timezone.now().strftime("%B %d, %Y %I:%M %p")}'])
        ws.append([])

        # Metrics
        metrics = [
            ('Metric', 'Value'),
            ('Total Patients', total_patients),
            ('New Patients (period)', new_patients_period),
            ('Gross Sales', float(gross_sales)),
            ('Net Sales', float(net_sales)),
            ('Total Discounts', float(total_discount)),
            ('Transactions', transaction_count),
            ('New Clients (period)', new_clients),
            ('Returning Clients (period)', returning_clients),
        ]

        for row_idx, (label, val) in enumerate(metrics, start=6):
            ws.cell(row=row_idx, column=1,
                    value=label).font = label_font if row_idx == 6 else value_font
            cell = ws.cell(row=row_idx, column=2, value=val)
            cell.font = label_font if row_idx == 6 else value_font
            if row_idx == 6:
                ws.cell(row=row_idx, column=1).fill = header_fill
                ws.cell(row=row_idx, column=1).font = header_font
                ws.cell(row=row_idx, column=2).fill = header_fill
                ws.cell(row=row_idx, column=2).font = header_font
            elif isinstance(val, float):
                cell.number_format = money_format
            ws.cell(row=row_idx, column=1).border = thin_border
            ws.cell(row=row_idx, column=2).border = thin_border

        # ── Sheet 2: Monthly Trends ───────────────────────────────────────
        ws2 = wb.create_sheet(f"Mon - {branch_safe}")
        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 18
        ws2.column_dimensions['E'].width = 18

        headers = ['Month', 'New Clients',
                   'Returning Clients', 'Gross Sales', 'Net Sales']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, m in enumerate(months_data, start=2):
            ws2.cell(row=row_idx, column=1, value=m['month']).border = thin_border
            ws2.cell(row=row_idx, column=2, value=m['new']).border = thin_border
            ws2.cell(row=row_idx, column=3,
                     value=m['returning']).border = thin_border
            gross_cell = ws2.cell(row=row_idx, column=4, value=m['gross'])
            gross_cell.number_format = money_format
            gross_cell.border = thin_border
            net_cell = ws2.cell(row=row_idx, column=5, value=m['net'])
            net_cell.number_format = money_format
            net_cell.border = thin_border

    if is_first_sheet:
        wb.active.title = "No Data"

    # ── Save and return ───────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'analytics_{period}_{date_from}_to_{date_to}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
